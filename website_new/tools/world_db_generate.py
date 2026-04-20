"""
Raldamain World Database Generator
===================================
Generates raldamain_db.xlsx with all world-building data organized in sheets.
Each sheet is a "table" with linked IDs for cross-referencing.

Usage: python tools/world_db_generate.py
Output: data/world/raldamain_db.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
OUTPUT_PATH = os.path.join(PROJECT_ROOT, "data", "world", "raldamain_db.xlsx")

# ── Styling ──
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Calibri", size=10)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
ID_FONT = Font(name="Consolas", size=10, color="8B4513")
LINK_FONT = Font(name="Consolas", size=9, color="2E75B6")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
STUB_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

# Category-specific header colors
SHEET_COLORS = {
    "Nations":    "2F5496",
    "Cultures":   "548235",
    "Cities":     "BF8F00",
    "Regions":    "7B5C3A",
    "Factions":   "C00000",
    "Characters": "7030A0",
    "Religions":  "2E75B6",
    "Deities":    "8B6914",
    "Items":      "385723",
    "Events":     "833C0B",
}


def style_sheet(ws, headers, data, color_hex=None):
    """Apply consistent styling to a worksheet."""
    fill_color = color_hex or "2F5496"
    hdr_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = hdr_fill
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value or "")
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER
            # ID column styling
            if col_idx == 1:
                cell.font = ID_FONT
            # Link columns (contain IDs referencing other sheets)
            if headers[col_idx - 1].endswith("_id") or headers[col_idx - 1].endswith("_ids"):
                cell.font = LINK_FONT
            # Alternate row shading
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL
            # Mark stub/incomplete rows (if last column says "STUB")
            if value == "STUB":
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=c).fill = STUB_FILL

    # Auto-width columns (capped)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_data in data:
            val = str(row_data[col_idx - 1] or "")
            max_len = max(max_len, min(len(val), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    # Freeze header row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"


# ═══════════════════════════════════════════════════════════════
# DATA DEFINITIONS
# ═══════════════════════════════════════════════════════════════

NATIONS_HEADERS = [
    "id", "name", "type", "capital_id", "government",
    "religion_id", "culture_id", "region", "status",
    "allies", "enemies", "notes",
]

NATIONS_DATA = [
    ("ustilus", "Imperio de Ustilus", "Empire", "trono_de_feyn",
     "Emperor + Parliament (Estatuto Lisandrino)", "feynismo;iglesia_heredero",
     "ustileses", "Central-North", "Active - Expansionist",
     "kratoi(tense)", "iskaria;temash_rebels",
     "Largest empire. Fileon is Emperor since 4010. Controlled by Culto de Sek-Tebos via Iglesia del Heredero."),
    ("mazarin", "Reino de Mazarín", "Kingdom", "svaikal",
     "Feudal monarchy (puppet king)", "feynismo",
     "mazarinos", "North", "Declining",
     "ustilus(nominal)", "",
     "Child king Netyalko VII. Controlled behind scenes by El Pacto. Defeated by Abshalom ~4003."),
    ("kratoi", "República de Kratoi", "Republic", "rimas",
     "Presidential democracy", "fe_espiritual",
     "kratenses", "Central-West", "Active - New",
     "mazarin", "ustilus(tense)",
     "First democracy. Founded 4025 by El Pacto. President Rezeda Valdei. Iglesia del Heredero infiltrating via Soucart."),
    ("iskaria", "Estado Reforjado de Iskaria", "Revolutionary State", "rajak_tai",
     "Military dictatorship (Comité de Acero)", "secular",
     "iskarianos", "East", "Active - Revolutionary",
     "", "elfos;ustilus",
     "Led by Shen Hyung. Anti-elf, anti-hobgoblin. Cutting Bosque Deju. Industrial revolution."),
    ("arshalan", "Imperio de Arshalán", "Empire", "beisyan",
     "Mhayid (religious-political leader)", "mekhatimismo",
     "leoninos;letemitas", "Southeast (Kimon)", "Active - Stable",
     "sindhal(Pacto de Sindhal);hakrapeu(vassal)", "",
     "Leonino ruling class + Letemita merchants. Founded ~3650 after Mekhatim's conversion of Canrim."),
    ("hakrapeu", "Hakrapeu", "Vassal State", "culthai",
     "Nine Noble Houses (Nueve Casas)", "cultos_abisales",
     "hakrapenses", "South (Kimon)", "Active - Vassal of Arshalán",
     "arshalan(suzerain)", "",
     "Mesoamerican-inspired swamp nation. Freed from dragon slavery by Nine Heroes. Worship Shade/Madre del Abismo."),
    ("thas_tapet", "Thas-Tapet", "Merchant Republic", "tapethavn",
     "Merchant council / Republic", "prisdevelianismo",
     "tapeties", "West-Central", "Active - Wealthy",
     "", "",
     "Dutch-inspired. Bolsa de Tapethavn (stock exchange). Pacto con Urlok. Tolerant of all cultures."),
    ("thiamashte", "Reino de Thiamashte", "Kingdom", "",
     "Elven monarchy", "fe_espiritual",
     "elfos", "East (Tierras Élficas)", "Active",
     "neronvain", "iskaria",
     "Conservative elven kingdom."),
    ("neronvain", "Federación de Neronvain", "Federation", "syralor",
     "Elven federation", "fe_espiritual",
     "elfos", "East (Tierras Élficas)", "Active",
     "thiamashte", "iskaria",
     "More open elven federation. Contains Syralor."),
    ("sindhal", "Sindhal", "Meritocracy", "el_pico",
     "Academic meritocracy", "secular",
     "aven", "Western Island", "Active",
     "arshalan(Pacto de Sindhal)", "",
     "Aven (avian humanoids) homeland. Rationalist atheist society."),
    ("tamashkhan", "Tamashkhan", "Conquered Territory", "kasheba",
     "Ustilus military occupation", "fe_ancestral_temash",
     "temash", "Central (Desert)", "Occupied by Ustilus",
     "", "ustilus",
     "Temash homeland. Conquered by Abshalom then Fileon. Rebels led by Rostam."),
]

CULTURES_HEADERS = [
    "id", "name", "inspiration", "nation_id", "religion_id",
    "language", "values", "key_customs", "notes",
]

CULTURES_DATA = [
    ("ustileses", "Ustileses", "Italian/Mediterranean", "ustilus", "feynismo",
     "Ustiliano", "Bella figura, honor, family, vendetta",
     "Opera; patrician families; hobgoblin slave soldiers",
     "Imperial culture in crisis after Feyn's disappearance."),
    ("temash", "Temash", "Berber/Tuareg", "tamashkhan", "fe_ancestral_temash",
     "Tamazight-inspired", "Asuf (honor code), tribal loyalty",
     "Tawsit clans; matrilineal; Polvo de Ángel (ritual drug); Tifinagh script",
     "Desert nomads. Oppressed by Ustilus. Rostam leads resistance."),
    ("kratenses", "Kratenses", "Celtic", "kratoi", "fe_espiritual",
     "Kratense", "Community, pragmatic idealism, freedom",
     "Fest-noz festivals; Biniou instrument; democratic participation",
     "Pragmatic idealists. First democratic culture."),
    ("letemitas", "Letemitas", "Semitic/Mediterranean", "arshalan", "mekhatimismo",
     "Letemita", "Commerce, banking, contracts, pragmatism",
     "Merchant guilds; family businesses; diaspora networks; detailed food culture",
     "Merchant class of Arshalán. Banking networks across Raldamain."),
    ("iskarianos", "Iskarianos", "Chinese/East Asian", "iskaria", "secular",
     "Iskariano", "Industry, collective labor, anti-colonialism",
     "Hwaro elemental spirits; steel/smoke aesthetic; secular ateísmo",
     "Descendants of Xiangdi. Hate elf colonialism and hobgoblin oppression."),
    ("leoninos", "Leoninos", "African/Feline", "arshalan", "mekhatimismo",
     "Leonino (Swahili-inspired)", "Warrior honor, strength, loyalty",
     "Kravah dueling; Regresión phenomenon (bestial relapse under stress)",
     "Feline humanoids. Warrior-noble class of Arshalán. Converted by Mekhatim."),
    ("elfos", "Elfos / Awen", "Celtic", "thiamashte;neronvain", "fe_espiritual",
     "Élfico (Celtic-inspired)", "Nature, spirits, Hiraeth (deep longing)",
     "Druidic hierarchy; Mundo Espiritual connection; formerly Imperio de Taranogh",
     "Immortal. Transformed humans via Awen. At war with Iskaria over Bosque Deju."),
    ("aven", "Aven", "Avian/Academic", "sindhal", "secular",
     "Aven", "Reason, empiricism, knowledge, meritocracy",
     "El Acantilado dueling; advancement through academic achievement",
     "Avian humanoids. Atheist rationalists."),
    ("hakrapenses", "Hakrapenses", "Mesoamerican", "hakrapeu", "cultos_abisales",
     "Hakrapense (Nahuatl/Mayan-insp.)", "Death, ancestors, liberation",
     "Nine Noble Houses; nigromancy; Shade worship; dragon-slavery liberation history",
     "Freed from dragon slavery by Nine Heroes."),
    ("tapeties", "Tapetíes", "Dutch/Flemish", "thas_tapet", "prisdevelianismo",
     "Tapetie", "Commerce, tolerance, pragmatism",
     "Bolsa de Tapethavn; Pacto con Urlok; tolerance of all cultures",
     "Merchant republic culture. Syncretic religion."),
    ("mazarinos", "Mazarinos", "Russian", "mazarin", "feynismo",
     "Mazarino", "Nobility, endurance, faith",
     "Feudal aristocracy; Estepa Blanca survival; Patriarchate in Svaikal",
     "STUB"),
    ("shirajianos", "Shirajianos", "Persian?", "", "culto_astral",
     "Shirajiano", "", "",
     "STUB"),
    ("shinri", "Shinri", "Japanese", "", "fe_espiritual",
     "Shinri", "Spirit harmony, tradition",
     "Encarnado legend; contested by Ustilus expansion",
     "STUB"),
]

CITIES_HEADERS = [
    "id", "name", "nation_id", "region", "type",
    "key_features", "notes",
]

CITIES_DATA = [
    ("trono_de_feyn", "Trono de Feyn", "ustilus", "Central-North", "Capital",
     "Massive capital; seat of Emperor and Iglesia del Heredero", ""),
    ("deverentio", "Deverentio", "ustilus", "Central", "Major City", "", "STUB"),
    ("elyon", "Elyon", "ustilus", "Central", "Major City", "", "STUB"),
    ("rizad", "Rizad", "ustilus", "Central", "Major City", "", "STUB"),
    ("noira", "Noira", "ustilus", "Thalos", "Provincial City", "Province of Thalos", ""),
    ("murat", "Murat", "ustilus", "Thalos", "Provincial City", "Province of Thalos", ""),
    ("svaikal", "Svaikal", "mazarin", "North", "Capital",
     "Seat of Feynismo Patriarchate (north)", ""),
    ("malekay", "Malekay", "mazarin", "North", "City", "", "STUB"),
    ("velotsk", "Velotsk", "mazarin", "North", "City", "", "STUB"),
    ("novotsar", "Novotsar", "mazarin", "North", "City", "", "STUB"),
    ("rimas", "Rimas", "kratoi", "West", "Capital",
     "Capital of the Republic", ""),
    ("argostana", "Argostana", "kratoi", "West", "Commercial Hub",
     "Commercial center; second city of Kratoi", ""),
    ("carpentia", "Carpentia", "kratoi", "Coast", "Naval Port",
     "Naval base and port city", ""),
    ("soucart", "Soucart", "kratoi", "West", "Wine Town",
     "Wine production; Iglesia del Heredero infiltrating here", "Threat vector"),
    ("bergemur", "Bergemur", "kratoi", "Coast", "Fishing Village",
     "Supernatural phenomena reported", ""),
    ("rajak_tai", "Rajak-Tai", "iskaria", "East", "Capital",
     "Capital of Estado Reforjado", ""),
    ("otamard", "Otamard", "iskaria", "East", "Industrial City",
     "Major industrial center", ""),
    ("beisyan", "Beisyan", "arshalan", "Southeast (Kimon)", "Major City",
     "Letemita primary city; commerce hub", ""),
    ("tapethavn", "Tapethavn", "thas_tapet", "West", "Capital",
     "Bolsa de Tapethavn (stock exchange)", ""),
    ("erevasai", "Erevasai", "thas_tapet", "West", "City", "", "STUB"),
    ("culthai", "Culthai", "hakrapeu", "South (Kimon)", "City", "", "STUB"),
    ("kasheba", "Kasheba", "tamashkhan", "Desert", "Port City",
     "Temash port; used by Abshalom during conquest", ""),
    ("lara_quayla", "Lara Quayla", "tamashkhan", "Desert", "City", "", "STUB"),
    ("menadar", "Menadar", "tamashkhan", "Desert", "City", "", "STUB"),
    ("qandesh", "Qandesh", "tamashkhan", "Desert", "City", "", "STUB"),
    ("syralor", "Syralor", "neronvain", "Tierras Élficas", "Elven City", "", "STUB"),
    ("el_pico", "El Pico", "sindhal", "Western Island", "Settlement",
     "Important Aven site; dueling arena El Acantilado", ""),
]

REGIONS_HEADERS = [
    "id", "name", "type", "nation_id",
    "description", "notes",
]

REGIONS_DATA = [
    ("thalos", "Thalos", "Province", "ustilus",
     "Contains Noira, Murat; Río Orzan; Bosque de Tobushi", ""),
    ("tamashkhan_r", "Tamashkhan (Region)", "Desert", "tamashkhan",
     "Temash desert homeland; conquered by Ustilus", ""),
    ("tierras_elficas", "Tierras Élficas", "Elven Lands", "thiamashte;neronvain",
     "Home of elves; split between Thiamashte and Neronvain", ""),
    ("estepa_blanca", "La Estepa Blanca", "Frozen Steppe", "mazarin",
     "Harsh frozen steppes of northern Mazarín", ""),
    ("yermos_eternos", "Yermos Eternos", "Wilderness", "",
     "Frozen north; Utopía/Nekthys ruins; Montañas Jaddi; Kobana; Tenebrio's domain", ""),
    ("brecha_abisal", "Brecha Abisal", "Abyss Rift", "",
     "Enormous chasm in the south; connection to the Abyss; Cultos Abisales", ""),
    ("kimon", "Kimon", "Subcontinent", "arshalan",
     "Arshalán subcontinent; southeast of Raldamain", ""),
    ("bosque_deju", "Bosque Deju", "Sacred Forest", "iskaria",
     "Sacred elven forest being cut down by Iskaria; active war zone", ""),
    ("canal_continental", "Canal Continental", "Waterway", "ustilus",
     "Massive artificial canal dividing the continent north/south; built by Ustilus", ""),
    ("tierras_shinri", "Tierras Shinri", "Region", "",
     "Shinri homeland; Encarnado legend; contested by Ustilus expansion", "STUB"),
    ("chothgar_mori", "Chothgar Mori", "Region", "",
     "Home of Camino de Yathan'ra cult", "STUB"),
    ("shiraj", "Shiraj", "Region", "",
     "Shirajianos homeland; Culto Astral", "STUB"),
    ("dairov", "Dairov", "Region", "",
     "Contains Pramanai, Lok Zigar", "STUB"),
    ("provincia_roumnet", "Provincia de Roumnet", "Province", "kratoi",
     "Dream magic practitioners; Plano Onírico connection; supernatural activity", ""),
    ("kajmasar", "Kajmasar", "Mountains", "",
     "Ancestral Leonino mountains", "STUB"),
    ("bosque_tobushi", "Bosque de Tobushi", "Forest", "ustilus",
     "Forest in Thalos province", "STUB"),
    ("montanas_jaddi", "Montañas Jaddi", "Mountains", "",
     "Mountains in Yermos Eternos; Abshalom's clan origin (Jaddi)", ""),
]

FACTIONS_HEADERS = [
    "id", "name", "type", "nation_id", "leader_id",
    "alignment", "goals", "status", "notes",
]

FACTIONS_DATA = [
    ("caballeros_ustilus", "Caballeros de Ustilus", "Military-Religious", "ustilus", "fileon",
     "Corrupted", "Serve Iglesia del Heredero; enforce imperial will",
     "Active - Corrupted",
     "Lost divine powers year 4000. Abshalom 4002-4005, Fileon 4005+. Now arm of Culto de Sek-Tebos."),
    ("orden_ascension", "Orden de la Ascensión", "Sub-order", "ustilus", "",
     "Corrupted", "Original holy sub-order", "Compromised", "Most affected by power loss."),
    ("diablos_plateados", "Diablos Plateados", "Sub-order", "ustilus", "",
     "Unclear", "Elite combat", "Active", ""),
    ("puro_fuego", "Puro Fuego", "Sub-order", "ustilus", "",
     "Inquisitorial", "Purification/inquisition", "Active", ""),
    ("hermandad_cuervo", "Hermandad del Cuervo", "Intelligence", "ustilus", "",
     "Unclear", "Intelligence gathering; espionage", "Active", "Spy sub-order."),
    ("orden_negra", "Orden Negra", "Sub-order", "ustilus", "",
     "Evil", "Unknown; fully under Sek-Tebos?", "Active", "Secretive."),
    ("el_pacto", "El Pacto", "Shadow Organization", "", "",
     "Neutral/Pragmatic", "Counterbalance Culto de Sek-Tebos; political control",
     "Active - Hidden",
     "Founded Kratoi. Controls Mazarín via puppet king. Opposes Iglesia del Heredero."),
    ("culto_sek_tebos", "Culto de Sek-Tebos", "Secret Cult", "", "basileos",
     "Evil", "Bind souls via Pacto del Servicio; give Sek-Tebos physical power",
     "Active - Hidden",
     "Operates through Iglesia del Heredero. Destroyed Rakashita theocracy historically. Only detectable by Mekhatimismo."),
    ("reforjados", "Reforjados", "Revolutionary Party", "iskaria", "shen_hyung",
     "Secular/Authoritarian", "Industrial revolution; anti-elf; anti-hobgoblin; secular equality",
     "Active - Ruling",
     "Won Guerra de las Cadenas Rotas 4006-4008. Comité de Acero governs. Oficina de la Llama = secret police."),
    ("legion_hobgoblin", "Legión Hobgoblin", "Military (Slave)", "ustilus", "",
     "Enslaved", "Serve as cannon fodder for Ustilus", "Active", "Hobgoblins enslaved by empire."),
    ("nueve_casas", "Nueve Casas", "Noble Houses", "hakrapeu", "",
     "Neutral", "Govern Hakrapeu; preserve Nine Heroes legacy", "Active",
     "Each house descends from one of the Nine Heroes."),
    ("escudos_rojos", "Escudos Rojos", "Unknown", "", "",
     "", "", "", "STUB"),
    ("zorros_negros", "Zorros Negros", "Unknown", "", "",
     "", "", "", "STUB"),
    ("clan_shikurai", "Clan Shikurai", "Clan", "", "",
     "", "", "", "STUB"),
]

CHARACTERS_HEADERS = [
    "id", "name", "title", "faction_id", "nation_id",
    "culture_id", "status", "role",
    "key_relationships", "notes",
]

CHARACTERS_DATA = [
    ("fileon", "Fileon", "Emperor / Gran Maestre", "caballeros_ustilus", "ustilus",
     "ustileses", "Alive", "Antagonist / Tragic Figure",
     "Abshalom(mentor); Basileos(manipulator); Ileria(wife,wounded/killed); Ajek(friend,lost)",
     "Gran Maestre 4005. Emperor 4010. Seduced by Basileos into Iglesia del Heredero. Tragic: believed in Ustilus but became puppet."),
    ("abshalom", "Abshalom", "Former Gran Maestre", "caballeros_ustilus", "ustilus",
     "ustileses", "Dead (4005)", "Historical Antagonist",
     "Fileon(mentee); Laicon(brother,betrayer); Ileria(complex); Yabu(ally); Madeleine(ally); Chandrial(ally)",
     "Jaddi clan, Taedros family. Coup 4002. Conquered Tamashkhan. Defeated Petrikov. Died at Mina de Katagral vs Rostam."),
    ("basileos", "Basileos", "Cult Leader", "culto_sek_tebos", "",
     "", "Alive", "Mastermind Villain",
     "Fileon(puppet); Sek-Tebos(master)",
     "Created Iglesia del Heredero. Mastermind of Culto de Sek-Tebos' current strategy."),
    ("shen_hyung", "Shen Hyung", "Mariscal Supremo", "reforjados", "iskaria",
     "iskarianos", "Alive", "Leader",
     "", "Led Guerra de las Cadenas Rotas. Rules Iskaria via Comité de Acero."),
    ("rezeda_valdei", "Rezeda Valdei", "President", "", "kratoi",
     "kratenses", "Alive", "Leader",
     "El Pacto(backers)", "First president of Kratoi. Elected via El Pacto's orchestration."),
    ("rostam", "Rostam", "Rebel Leader", "", "tamashkhan",
     "temash", "Alive?", "Resistance Hero",
     "Abshalom(nemesis)", "Led Temash rebels. Cornered Abshalom at Mina de Katagral."),
    ("netyalko_vii", "Netyalko VII", "King (Child)", "", "mazarin",
     "mazarinos", "Alive", "Puppet Ruler",
     "El Pacto(controllers)", "Child king. Puppet of El Pacto."),
    ("ileria", "Ileria", "", "caballeros_ustilus", "ustilus",
     "ustileses", "Dead/Wounded?", "Tragic Figure",
     "Fileon(husband); Abshalom(complex)", "Wounded or killed by Fileon."),
    ("laicon", "Laicon", "", "", "ustilus",
     "ustileses", "Unknown", "Betrayer",
     "Abshalom(brother-figure,betrayed)", "Betrayed Abshalom."),
    ("acrodio", "Acrodio / Akrodios", "Founder", "caballeros_ustilus", "ustilus",
     "ustileses", "Dead (Historical)", "Founder",
     "Feyn(received prophecies)", "Founded Ustilus and Caballeros in 3496. Received prophecy of Feyn's return in 4000."),
    ("tenebrio", "Tenebrio / Alithei", "Emperor", "", "",
     "", "Undead", "Historical Villain",
     "Feyn(rebelled against); Ayanuu(driven mad by)",
     "Emperor ~2207. Rebelled against Feyn. Defeated. Banished to undead existence in Yermos Eternos."),
    ("mekhatim", "Mekhatim", "Prophet", "", "",
     "", "Dead (Historical)", "Founder",
     "Ithanatron(received Taurenet); Canrim(converted)",
     "Founded Mekhatimismo in 3646. Received Taurenet from Celestial Ithanatron."),
    ("canrim", "Canrim", "First Mhayid", "", "arshalan",
     "leoninos", "Dead (Historical)", "Founder",
     "Mekhatim(converted by)", "First Mhayid of Arshalán. Leonino chief converted by Mekhatim."),
    ("encarnado", "Encarnado", "Legendary Hero", "", "",
     "shinri", "Reincarnating", "Mythic Hero",
     "Ekdavor(hunted by)", "Hero of Shinri lands. Reincarnates. Hunted by Abisal Ekdavor across lives."),
    ("zarin_i", "Zarin I", "Prince/Founder", "", "mazarin",
     "mazarinos", "Dead (Historical)", "Founder",
     "", "Founded Reino de Mazarín in 2481."),
    ("petrikov", "Petrikov", "General", "", "mazarin",
     "mazarinos", "Dead?", "Military",
     "Abshalom(defeated by)", "Mazarín general defeated by Abshalom ~4003."),
    ("ajek", "Ajek", "", "", "ustilus",
     "ustileses", "Dead?", "Friend of Fileon",
     "Fileon(close friend,lost)", "STUB"),
    ("yabu", "Yabu", "", "", "",
     "", "Unknown", "Ally of Abshalom",
     "Abshalom(ally)", "STUB"),
    ("madeleine", "Madeleine", "", "", "",
     "", "Unknown", "Ally of Abshalom",
     "Abshalom(ally/friend)", "STUB"),
    ("chandrial", "Chandrial", "", "", "",
     "", "Unknown", "Unknown",
     "Abshalom(mentioned)", "STUB"),
    ("meisodh", "Meisodh el Viajero", "Prophet/Warlord", "", "",
     "", "Dead (Historical)", "Cult Founder",
     "", "Spread Camino de Yathan'ra to Chothgar Mori. STUB"),
]

RELIGIONS_HEADERS = [
    "id", "name", "type", "deity_ids", "sacred_text",
    "origin", "strongholds", "notes",
]

RELIGIONS_DATA = [
    ("feynismo", "Feynismo", "Monotheist", "feyn", "",
     "Founded by Feyn year 1 D.F.", "ustilus;mazarin",
     "Dominant. In crisis since Feyn's disappearance (4000). Patriarchate in Svaikal. Being replaced by Iglesia del Heredero in Ustilus."),
    ("iglesia_heredero", "Iglesia del Heredero", "Fake/Cult", "sek_tebos",
     "Pacto del Servicio (soul contract)", "Created ~4006 by Basileos",
     "ustilus;soucart(kratoi)",
     "FAKE religion. Front for Culto de Sek-Tebos. Arcontes are demons. Fileon is figurehead. Binds souls."),
    ("mekhatimismo", "Mekhatimismo", "Monotheist", "hueste_celestial;ithanatron",
     "Taurenet", "Founded 3646 by Mekhatim",
     "arshalan;beisyan;letemita_diaspora",
     "Universalist. Only religion that can detect Culto de Sek-Tebos. Celestial hierarchy worship."),
    ("fe_espiritual", "Fe Espiritual", "Animist", "senores_mundo_espiritual",
     "", "Oldest religion; predates all others",
     "tierras_elficas;shinri;kratoi(rural);mazarin(rural)",
     "Animist. Señores del Mundo Espiritual. Elves transformed through this. Druidic. No central hierarchy."),
    ("cultos_abisales", "Cultos Abisales", "Nihilist Cults", "ayanuu;ekdavor;shade;gulmaht",
     "", "Ancient",
     "hakrapeu;chothgar_mori",
     "Four Abisales. Nihilistic: world is broken. Openly practiced only in Hakrapeu and Chothgar Mori."),
    ("prisdevelianismo", "Prisdevelianismo", "Syncretic", "",
     "", "Tapetíes syncretic creation",
     "thas_tapet",
     "Combines Feynismo + Fe Espiritual + merchant philosophy. Pacto con Urlok. Pragmatic."),
    ("fe_ancestral_temash", "Fe Ancestral Temash", "Ancestor Worship", "",
     "", "Temash traditional",
     "tamashkhan",
     "Tied to Asuf honor code and ancestor veneration."),
    ("culto_astral", "Culto Astral", "Astronomical", "",
     "", "Shirajiano origin",
     "shiraj",
     "Star/astronomical worship. STUB"),
    ("camino_yathanra", "Camino de Yathan'ra", "Abisal Cult", "",
     "", "Spread by Meisodh el Viajero",
     "chothgar_mori",
     "Specific Abisal cult. STUB"),
]

DEITIES_HEADERS = [
    "id", "name", "type", "domain", "religion_id",
    "status", "notes",
]

DEITIES_DATA = [
    ("feyn", "Feyn", "God", "Humanity, civilization, law", "feynismo",
     "Disappeared (4000 D.F.)",
     "Descended year 1. Founded República Feynista. Gave Acrodio prophecy. Disappeared year 4000 — no explanation."),
    ("sek_tebos", "Sek-Tebos", "Fallen Celestial", "Domination, corruption, soul-binding", "culto_sek_tebos",
     "Active (hidden)",
     "Not originally a god. Fallen Celestial. Seeks physical power through bound souls via Pacto del Servicio."),
    ("ithanatron", "Ithanatron", "Celestial", "Divine revelation", "mekhatimismo",
     "Active",
     "Gave Mekhatim the Taurenet. Higher celestial of the Hueste Celestial."),
    ("ayanuu", "Ayanuu", "Abisal", "Nihilistic destruction, madness", "cultos_abisales",
     "Active",
     "Linked to Tenebrio's madness. War against Nekthys/Utopía."),
    ("ekdavor", "Ekdavor", "Abisal", "Vengeance, soul-hunting", "cultos_abisales",
     "Active",
     "Hunts the Encarnado across reincarnations."),
    ("shade", "Shade / Madre del Abismo", "Abisal", "Death, swamp, underworld", "cultos_abisales",
     "Active",
     "Worshipped in Hakrapeu. Nine Heroes guided by her."),
    ("gulmaht", "Gulmaht", "Abisal", "War, conquest", "cultos_abisales",
     "Active", ""),
    ("hueste_celestial", "Hueste Celestial", "Celestial Host", "Divine order", "mekhatimismo",
     "Active", "Collective angelic hierarchy. Supreme deity above them."),
    ("senores_mundo_espiritual", "Señores del Mundo Espiritual", "Spirit Lords", "Nature, spirits", "fe_espiritual",
     "Active", "Lords of the Spirit World. Animist powers. Tied to Primeros Dioses."),
    ("primeros_dioses", "Primeros Dioses", "Primordial", "Creation, origin", "",
     "Historical/Unknown", "Progenitors of all divine powers. Guerra de los Dioses."),
    ("chiso_shitari", "Chiso-Shitari", "Spirit?", "", "fe_espiritual",
     "", "STUB"),
    ("seiryuu", "Seiryuu", "Spirit/Dragon?", "Azure Dragon?", "fe_espiritual",
     "", "STUB - Japanese Azure Dragon association"),
    ("suzakushin", "Suzakushin", "Spirit?", "Vermilion Bird?", "fe_espiritual",
     "", "STUB"),
    ("zornovog", "Zornovog", "Unknown", "", "",
     "", "STUB - Slavic name; possibly Mazarín deity"),
    ("la_tormenta", "La Tormenta", "Primordial?", "Storm", "",
     "", "STUB"),
]

ITEMS_HEADERS = [
    "id", "name", "type", "origin",
    "significance", "notes",
]

ITEMS_DATA = [
    ("taurenet", "Taurenet", "Sacred Text", "Given by Ithanatron to Mekhatim",
     "Foundation of Mekhatimismo. Contains divine laws and theology.",
     ""),
    ("chi_crystals", "Cristales de Chi", "Magical Resource", "Ustilus",
     "Key economic resource. Used in industry and magic. Powers the Ustilese economy.",
     ""),
    ("polvo_de_angel", "Polvo de Ángel", "Ritual Drug", "Temash",
     "Used in Temash rituals and recreationally. Cultural significance.",
     ""),
    ("pacto_servicio", "Pacto del Servicio", "Soul Contract", "Culto de Sek-Tebos",
     "Soul-binding contract given to Iglesia del Heredero initiates. Traps souls for Sek-Tebos.",
     "The primary tool of Culto de Sek-Tebos."),
]

EVENTS_HEADERS = [
    "id", "name", "year", "nation_ids",
    "character_ids", "description", "consequences",
]

EVENTS_DATA = [
    ("feyn_descends", "Feyn Descends", "1 D.F.", "",
     "feyn", "Feyn descends to Raldamain; founds the República Feynista",
     "Beginning of recorded history. Humanity receives civilization."),
    ("tenebrio_rebellion", "Rebelión de Tenebrio", "~2207", "",
     "tenebrio;feyn", "Emperor Tenebrio (Alithei) rebels against Feyn; driven mad by Ayanuu",
     "Tenebrio defeated and banished to undead existence in Yermos Eternos."),
    ("mazarin_founded", "Fundación de Mazarín", "2481", "mazarin",
     "zarin_i", "Prince Zarin I founds the Reino de Mazarín",
     "Feudal monarchy established in the north."),
    ("ustilus_founded", "Fundación del Imperio de Ustilus", "3496", "ustilus",
     "acrodio", "Acrodio founds Ustilus and Caballeros; receives Feyn's prophecy of year 4000",
     "Largest empire begins. Acrodio builds toward Feyn's prophesied return."),
    ("mekhatim_ministry", "Ministerio de Mekhatim", "3646", "arshalan",
     "mekhatim;ithanatron;canrim",
     "Mekhatim receives Taurenet; converts Canrim; founds Mekhatimismo",
     "Arshalán founded. New monotheistic religion spreads."),
    ("feyn_disappears", "La Desaparición de Feyn", "4000", "ustilus;mazarin",
     "feyn", "Feyn simply disappears without explanation",
     "Era de la Incertidumbre begins. Caballeros lose divine powers. Religious vacuum exploited by Culto de Sek-Tebos."),
    ("abshalom_coup", "Golpe de Abshalom", "4002", "ustilus",
     "abshalom", "Abshalom seizes control of Ustilus as Gran Maestre",
     "Military dictatorship. Campaigns in Mazarín and Tamashkhan."),
    ("abshalom_vs_petrikov", "Derrota de Petrikov", "~4003", "mazarin;ustilus",
     "abshalom;petrikov", "Abshalom defeats Mazarín general Petrikov",
     "Mazarín weakened. El Pacto gains more control behind scenes."),
    ("abshalom_death", "Muerte de Abshalom", "~4005", "ustilus;tamashkhan",
     "abshalom;rostam", "Temash rebels led by Rostam corner Abshalom at Mina de Katagral; he dies",
     "Fileon becomes Gran Maestre. Tamashkhan occupation continues."),
    ("iglesia_heredero_created", "Creación de Iglesia del Heredero", "~4006", "ustilus",
     "basileos;fileon", "Basileos approaches Fileon; creates Iglesia del Heredero as front for Culto de Sek-Tebos",
     "Ustilus begins religious transformation. Soul-binding Pacto del Servicio spreads."),
    ("guerra_cadenas_rotas", "Guerra de las Cadenas Rotas", "4006-4008", "iskaria",
     "shen_hyung", "Reforjados party leads revolution in Iskaria",
     "Estado Reforjado de Iskaria established. Shen Hyung rules. Anti-elf, industrial."),
    ("fileon_emperor", "Fileon se Proclama Emperador", "4010", "ustilus",
     "fileon", "Fileon becomes Emperor of Ustilus — political and religious supreme authority",
     "Ustilus fully under Iglesia del Heredero control. Expansion toward Shinri."),
    ("kratoi_founded", "Fundación de Kratoi", "4025", "kratoi",
     "rezeda_valdei", "República de Kratoi founded — first democracy. Orchestrated by El Pacto",
     "Constitution ratified 4027. Independence from Ustilus."),
]


def main():
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    sheets_config = [
        ("Nations", NATIONS_HEADERS, NATIONS_DATA),
        ("Cultures", CULTURES_HEADERS, CULTURES_DATA),
        ("Cities", CITIES_HEADERS, CITIES_DATA),
        ("Regions", REGIONS_HEADERS, REGIONS_DATA),
        ("Factions", FACTIONS_HEADERS, FACTIONS_DATA),
        ("Characters", CHARACTERS_HEADERS, CHARACTERS_DATA),
        ("Religions", RELIGIONS_HEADERS, RELIGIONS_DATA),
        ("Deities", DEITIES_HEADERS, DEITIES_DATA),
        ("Items", ITEMS_HEADERS, ITEMS_DATA),
        ("Events", EVENTS_HEADERS, EVENTS_DATA),
    ]

    for sheet_name, headers, data in sheets_config:
        ws = wb.create_sheet(title=sheet_name)
        color = SHEET_COLORS.get(sheet_name, "2F5496")
        style_sheet(ws, headers, data, color)
        ws.sheet_properties.tabColor = color

    # Save
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Database generated: {OUTPUT_PATH}")
    print(f"Sheets: {', '.join(s[0] for s in sheets_config)}")
    print(f"Total entries: {sum(len(s[2]) for s in sheets_config)}")


if __name__ == "__main__":
    main()

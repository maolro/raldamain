"""
Raldamain World Database Reader
================================
Reads raldamain_db.xlsx and provides query/lookup utilities.
Can be used as a library (import) or CLI tool.

Usage:
  python tools/world_db_reader.py                     # Print summary
  python tools/world_db_reader.py list <sheet>        # List all entries in a sheet
  python tools/world_db_reader.py get <sheet> <id>    # Get one entry by ID
  python tools/world_db_reader.py search <term>       # Search all sheets for a term
  python tools/world_db_reader.py links <sheet> <id>  # Show cross-references for an entry
  python tools/world_db_reader.py stubs               # List all STUB entries
  python tools/world_db_reader.py export <format>     # Export to json or csv

Sheets: Nations, Cultures, Cities, Regions, Factions, Characters, Religions, Deities, Items, Events
"""

import os
import sys
import json
import csv
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DB_PATH = os.path.join(PROJECT_ROOT, "data", "world", "raldamain_db.xlsx")

# Columns that contain IDs referencing other sheets
LINK_COLUMNS = {
    "capital_id": "Cities",
    "nation_id": "Nations",
    "nation_ids": "Nations",
    "culture_id": "Cultures",
    "religion_id": "Religions",
    "deity_ids": "Deities",
    "leader_id": "Characters",
    "faction_id": "Factions",
    "character_ids": "Characters",
}


class WorldDB:
    """In-memory representation of the world database."""

    def __init__(self, path=None):
        self.path = path or DB_PATH
        self._sheets = {}
        self._load()

    def _load(self):
        wb = load_workbook(self.path, read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h or "").strip() for h in rows[0]]
            entries = []
            for row in rows[1:]:
                entry = {}
                for i, header in enumerate(headers):
                    val = row[i] if i < len(row) else ""
                    entry[header] = str(val) if val else ""
                entries.append(entry)
            self._sheets[name] = {"headers": headers, "entries": entries}
        wb.close()

    @property
    def sheet_names(self):
        return list(self._sheets.keys())

    def get_sheet(self, name):
        """Get all entries from a sheet. Case-insensitive match."""
        for key in self._sheets:
            if key.lower() == name.lower():
                return self._sheets[key]["entries"]
        return None

    def get_headers(self, name):
        """Get headers for a sheet."""
        for key in self._sheets:
            if key.lower() == name.lower():
                return self._sheets[key]["headers"]
        return None

    def get_by_id(self, sheet_name, entry_id):
        """Get a single entry by its ID (first column)."""
        entries = self.get_sheet(sheet_name)
        if not entries:
            return None
        for entry in entries:
            if entry.get("id", "") == entry_id:
                return entry
        return None

    def search(self, term, sheet_name=None):
        """Search all sheets (or one) for a term. Returns list of (sheet, entry)."""
        results = []
        term_lower = term.lower()
        sheets = [sheet_name] if sheet_name else self.sheet_names
        for name in sheets:
            entries = self.get_sheet(name)
            if not entries:
                continue
            for entry in entries:
                for val in entry.values():
                    if term_lower in val.lower():
                        results.append((name, entry))
                        break
        return results

    def get_stubs(self):
        """Get all entries that contain STUB markers."""
        stubs = []
        for name in self.sheet_names:
            for entry in self.get_sheet(name):
                for val in entry.values():
                    if "STUB" in val.upper():
                        stubs.append((name, entry))
                        break
        return stubs

    def get_links(self, sheet_name, entry_id):
        """Find all cross-references for an entry.
        Returns dict of {target_sheet: [linked_entries]}."""
        entry = self.get_by_id(sheet_name, entry_id)
        if not entry:
            return {}

        links = {}

        # Forward links: columns in this entry that reference other sheets
        for col, target_sheet in LINK_COLUMNS.items():
            if col in entry and entry[col]:
                ids = [x.strip() for x in entry[col].split(";") if x.strip()]
                for ref_id in ids:
                    # Strip parenthetical notes like "ustilus(nominal)"
                    clean_id = ref_id.split("(")[0].strip()
                    target = self.get_by_id(target_sheet, clean_id)
                    if target:
                        links.setdefault(target_sheet, []).append(target)

        # Reverse links: entries in other sheets that reference this entry's ID
        for other_sheet in self.sheet_names:
            headers = self.get_headers(other_sheet)
            if not headers:
                continue
            for col in headers:
                if col in LINK_COLUMNS or col.endswith("_id") or col.endswith("_ids"):
                    for other_entry in self.get_sheet(other_sheet):
                        val = other_entry.get(col, "")
                        ids = [x.split("(")[0].strip() for x in val.split(";")]
                        if entry_id in ids:
                            key = f"{other_sheet} (reverse)"
                            links.setdefault(key, []).append(other_entry)

        return links

    def summary(self):
        """Return a summary dict of sheet names and entry counts."""
        return {name: len(data["entries"]) for name, data in self._sheets.items()}

    def export_json(self, output_dir=None):
        """Export all sheets as individual JSON files."""
        out = output_dir or os.path.join(PROJECT_ROOT, "data", "world", "export")
        os.makedirs(out, exist_ok=True)
        for name in self.sheet_names:
            entries = self.get_sheet(name)
            path = os.path.join(out, f"{name.lower()}.json")
            with open(path, "w", encoding="utf-8") as f:
                json.dump(entries, f, indent=2, ensure_ascii=False)
        return out

    def export_csv(self, output_dir=None):
        """Export all sheets as individual CSV files."""
        out = output_dir or os.path.join(PROJECT_ROOT, "data", "world", "export")
        os.makedirs(out, exist_ok=True)
        for name in self.sheet_names:
            headers = self.get_headers(name)
            entries = self.get_sheet(name)
            path = os.path.join(out, f"{name.lower()}.csv")
            with open(path, "w", encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(entries)
        return out


# ── CLI ──

def _print_entry(entry, indent=0):
    pad = "  " * indent
    for k, v in entry.items():
        if v:
            print(f"{pad}{k}: {v}")


def _print_table(headers, entries, max_col_width=40):
    """Print entries as a formatted table."""
    # Use only first 5 columns for compact display
    show = headers[:5]
    widths = {h: len(h) for h in show}
    for entry in entries:
        for h in show:
            widths[h] = min(max(widths[h], len(entry.get(h, ""))), max_col_width)

    # Header
    row = " | ".join(h.ljust(widths[h]) for h in show)
    print(row)
    print("-+-".join("-" * widths[h] for h in show))

    # Rows
    for entry in entries:
        vals = []
        for h in show:
            v = entry.get(h, "")
            if len(v) > max_col_width:
                v = v[: max_col_width - 3] + "..."
            vals.append(v.ljust(widths[h]))
        print(" | ".join(vals))


def main():
    if not os.path.exists(DB_PATH):
        print(f"Database not found at {DB_PATH}")
        print("Run: python tools/world_db_generate.py")
        sys.exit(1)

    db = WorldDB()

    if len(sys.argv) < 2:
        # Summary
        print("=== Raldamain World Database ===")
        print(f"Source: {DB_PATH}\n")
        total = 0
        for name, count in db.summary().items():
            stubs = sum(1 for e in db.get_sheet(name) for v in e.values() if "STUB" in v.upper())
            total += count
            stub_str = f"  ({stubs} stubs)" if stubs else ""
            print(f"  {name:15s} {count:3d} entries{stub_str}")
        print(f"\n  {'TOTAL':15s} {total:3d} entries")
        return

    cmd = sys.argv[1].lower()

    if cmd == "list":
        if len(sys.argv) < 3:
            print(f"Usage: {sys.argv[0]} list <sheet>")
            print(f"Sheets: {', '.join(db.sheet_names)}")
            return
        sheet = sys.argv[2]
        entries = db.get_sheet(sheet)
        headers = db.get_headers(sheet)
        if entries is None:
            print(f"Sheet '{sheet}' not found. Available: {', '.join(db.sheet_names)}")
            return
        _print_table(headers, entries)

    elif cmd == "get":
        if len(sys.argv) < 4:
            print(f"Usage: {sys.argv[0]} get <sheet> <id>")
            return
        sheet, eid = sys.argv[2], sys.argv[3]
        entry = db.get_by_id(sheet, eid)
        if entry:
            _print_entry(entry)
        else:
            print(f"No entry '{eid}' in sheet '{sheet}'")

    elif cmd == "search":
        if len(sys.argv) < 3:
            print(f"Usage: {sys.argv[0]} search <term>")
            return
        term = " ".join(sys.argv[2:])
        results = db.search(term)
        if not results:
            print(f"No results for '{term}'")
            return
        print(f"Found {len(results)} result(s) for '{term}':\n")
        for sheet, entry in results:
            print(f"[{sheet}] {entry.get('id', '?')} — {entry.get('name', entry.get('id', ''))}")
            _print_entry(entry, indent=1)
            print()

    elif cmd == "links":
        if len(sys.argv) < 4:
            print(f"Usage: {sys.argv[0]} links <sheet> <id>")
            return
        sheet, eid = sys.argv[2], sys.argv[3]
        links = db.get_links(sheet, eid)
        if not links:
            print(f"No links found for '{eid}' in '{sheet}'")
            return
        entry = db.get_by_id(sheet, eid)
        print(f"Links for [{sheet}] {entry.get('name', eid)}:\n")
        for target, entries in links.items():
            print(f"  -> {target}:")
            for e in entries:
                print(f"    * {e.get('id', '?')} - {e.get('name', e.get('id', ''))}")
        print()

    elif cmd == "stubs":
        stubs = db.get_stubs()
        if not stubs:
            print("No STUB entries found!")
            return
        print(f"Found {len(stubs)} STUB entries:\n")
        for sheet, entry in stubs:
            print(f"  [{sheet}] {entry.get('id', '?')} — {entry.get('name', entry.get('id', ''))}")

    elif cmd == "export":
        fmt = sys.argv[2].lower() if len(sys.argv) > 2 else "json"
        if fmt == "json":
            out = db.export_json()
            print(f"Exported JSON to: {out}")
        elif fmt == "csv":
            out = db.export_csv()
            print(f"Exported CSV to: {out}")
        else:
            print(f"Unknown format '{fmt}'. Use 'json' or 'csv'.")

    else:
        print(f"Unknown command '{cmd}'.")
        print("Commands: list, get, search, links, stubs, export")


if __name__ == "__main__":
    main()
